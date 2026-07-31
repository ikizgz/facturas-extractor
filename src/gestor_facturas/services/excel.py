"""Módulo de gestión y volcado de datos a archivos Excel usando openpyxl."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from gestor_facturas.core.config import ARCHIVO_EXCEL_SALIDA

# Logger específico para este módulo
logger = logging.getLogger(__name__)


def _inicializar_o_cargar_workbook(
    excel_path: str | Path,
) -> tuple[openpyxl.Workbook, Worksheet]:
    """Carga un workbook existente o crea uno nuevo con su cabecera estándar."""
    path = Path(excel_path)

    if path.exists():
        logger.info(f"Cargando archivo Excel existente: {path.name}")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title="Facturas")
    else:
        logger.info(f"Creando un nuevo archivo Excel: {path.name}")
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Facturas"
            # Cabeceras estándar actualizadas según los requerimientos
            cabeceras = [
                "Fecha",
                "Nº Factura",
                "Proveedor",
                "NIF/CIF",
                "Importe Base",
                "Tasa % IVA",
                "Importe IVA",
                "Importe Total",
            ]
            ws.append(cabeceras)
        else:
            raise RuntimeError("No se pudo inicializar la hoja activa del workbook.")

    return wb, ws


def guardar_factura_excel(
    datos_factura: dict,
    excel_path: str | Path = ARCHIVO_EXCEL_SALIDA,
) -> bool:
    """Añade los datos extraídos de una factura como una nueva fila en el archivo Excel.

    Espera un diccionario con las claves correspondientes a los nuevos campos.
    """
    path = Path(excel_path)

    try:
        wb, ws = _inicializar_o_cargar_workbook(path)

        # Extraer los campos en el orden exacto de las nuevas cabeceras
        fila = [
            datos_factura.get("fecha", ""),
            datos_factura.get("numero_factura", ""),
            datos_factura.get("proveedor", "Desconocido"),
            datos_factura.get("nif_cif", ""),
            datos_factura.get("importe_base", 0.0),
            datos_factura.get("tasa_iva", 0.0),
            datos_factura.get("importe_iva", 0.0),
            datos_factura.get("importe_total", 0.0),
        ]

        ws.append(fila)
        wb.save(path)
        logger.info(
            f"Factura de '{datos_factura.get('proveedor')}' ({datos_factura.get('numero_factura')}) "
            f"guardada correctamente en {path.name}"
        )
        return True

    except (OSError, PermissionError) as e:
        logger.error(
            "Error de permisos o E/S al guardar en el archivo Excel %s: %s",
            path.name,
            e,
        )
        return False
    except Exception:
        logger.exception(
            "Error inesperado al intentar guardar los datos en el Excel %s",
            path.name,
        )
        return False

# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import gc
import json
import logging
from pathlib import Path
from typing import Dict, List

import pandas as pd
from PyPDF2 import PdfReader

from providers import PROVIDERS
from providers.common import Row

DEFAULT_DPI = 150
DEFAULT_SLEEP_MS = 0
DEFAULT_THROTTLE_EVERY = 6
DEFAULT_THROTTLE_MS = 800
DEFAULT_CHILD_TIMEOUT_S = 60


# ---- OCR ----
def read_pdf_text_native(path: Path) -> str:
    try:
        reader = PdfReader(str(path))
        return "\n".join([(p.extract_text() or "") for p in reader.pages]) or ""
    except Exception:
        return ""


def read_pdf_text_ocr(
    path: Path, dpi: int, sleep_ms: int, poppler_path: str, tesseract_exe: str
) -> str:
    try:
        import pytesseract
        from pdf2image import convert_from_path, pdfinfo_from_path

        if tesseract_exe:
            pytesseract.pytesseract.tesseract_cmd = tesseract_exe
        info = pdfinfo_from_path(str(path), userpw="", poppler_path=poppler_path)
        max_pages = int(info.get("Pages", 1))
        collected: List[str] = []
        for page in range(1, max_pages + 1):
            images = convert_from_path(
                str(path),
                dpi=dpi,
                first_page=page,
                last_page=page,
                fmt="jpeg",
                thread_count=1,
                poppler_path=poppler_path,
            )
            img = images[0]
            cfg_text = "--psm 6"
            collected.append(
                pytesseract.image_to_string(img, lang="spa+eng", config=cfg_text)
            )
            try:
                img.close()
            except Exception:
                pass
            del images
            gc.collect()
            if sleep_ms > 0:
                import time

                time.sleep(sleep_ms / 1000.0)
        return "\n".join(collected)
    except Exception:
        return ""


# ---- Child isolation ----
def child_worker(
    q, pdf_path: str, dpi: int, sleep_ms: int, poppler_path: str, tesseract_exe: str
):
    try:
        path = Path(pdf_path)
        text = read_pdf_text_native(path)
        ocr_used = False
        if len(text) < 80:
            text = read_pdf_text_ocr(
                path,
                dpi=dpi,
                sleep_ms=sleep_ms,
                poppler_path=poppler_path,
                tesseract_exe=tesseract_exe,
            )
            ocr_used = True
        rows: List[Row] = []
        # Detección por proveedor (primero que detecte)
        for parser in PROVIDERS:
            if parser.detect(text):
                parsed_rows = parser.parse(text, path)
                # Añadir nota OCR
                for r in parsed_rows:
                    if ocr_used:
                        r["Notas"] = (
                            (r.get("Notas") or "")
                            + ("; " if r.get("Notas") else "")
                            + "OCR"
                        )
                rows.extend(parsed_rows)
                break
        if not rows:
            rows = [{"numero_factura": path.stem, "Notas": "Sin parser"}]
        q.put(json.dumps(rows))
    except Exception as e:
        p = Path(pdf_path)
        q.put(json.dumps([{"numero_factura": p.name, "Notas": f"Error: {e}"}]))


def run_child_extract(
    pdf_path: Path,
    dpi: int,
    sleep_ms: int,
    poppler_path: str,
    tesseract_exe: str,
    timeout_s: int,
) -> List[Row]:
    import multiprocessing as mp
    from multiprocessing import Process, Queue

    q: Queue = mp.Queue(maxsize=1)
    p = Process(
        target=child_worker,
        args=(q, str(pdf_path), dpi, sleep_ms, poppler_path, tesseract_exe),
        daemon=True,
    )
    p.start()
    p.join(timeout_s)
    if p.is_alive():
        try:
            p.terminate()
            p.join(5)
        except Exception:
            pass
        return [{"numero_factura": pdf_path.name, "Notas": f"Timeout {timeout_s}s"}]
    try:
        data = q.get_nowait()
        return json.loads(data)
    except Exception:
        return [{"numero_factura": pdf_path.name, "Notas": "Sin datos del hijo"}]


# ---- Excel formatting ----
def format_excel(path_out: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.worksheet import Worksheet
    except Exception:
        return
    try:
        wb = load_workbook(str(path_out))
    except Exception:
        return
    if not wb.sheetnames:
        return
    ws = wb.active
    # Harden for static analyzers (Pylance): ensure ws is Worksheet
    if not isinstance(ws, Worksheet):
        wb.save(str(path_out))
        return
    # Header row
    try:
        header_row = list(ws[1])
    except Exception:
        wb.save(str(path_out))
        return
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(header_row, start=1):
        key = str(cell.value).strip() if cell.value is not None else f"COL_{idx}"
        headers[key] = idx
    fmt_fecha = "dd/mm/yyyy"
    fmt_moneda = '"€"#,##0.00'
    fmt_pct = "0.00%"
    max_row = ws.max_row if hasattr(ws, "max_row") else 0
    if not isinstance(max_row, int) or max_row < 2:
        wb.save(str(path_out))
        return
    col_fecha = headers.get("fecha_factura")
    col_base = headers.get("importe_base")
    col_iva = headers.get("IVA")
    col_total = headers.get("importe_total")
    col_pct = headers.get("%IVA")
    if isinstance(col_fecha, int):
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col_fecha).number_format = fmt_fecha
    for col in (col_base, col_iva, col_total):
        if isinstance(col, int):
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=col).number_format = fmt_moneda
    if isinstance(col_pct, int):
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col_pct).number_format = fmt_pct
    try:
        wb.save(str(path_out))
    except Exception:
        pass


# ---- MAIN ----
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extractor modular PDF → Excel (parsers por proveedor)"
    )
    parser.add_argument(
        "--input", "-i", type=str, required=True, help="Carpeta con los PDFs"
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        default=None,
        help="Excel de salida (por defecto: facturas_datos_extraidos.xlsx)",
    )
    parser.add_argument(
        "--ocr",
        choices=["on", "off"],
        default="on",
        help="OCR si el PDF no tiene texto",
    )
    parser.add_argument(
        "--dpi",
        type=int,
        default=DEFAULT_DPI,
        help=f"DPI para OCR (por defecto {DEFAULT_DPI})",
    )
    parser.add_argument(
        "--poppler",
        type=str,
        default="",
        help="Ruta a binarios de Poppler si no están en PATH",
    )
    parser.add_argument(
        "--tesseract",
        type=str,
        default="",
        help="Ruta completa a tesseract si no está en PATH",
    )
    parser.add_argument(
        "--log", type=str, default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"]
    )
    parser.add_argument(
        "--sleep-ms",
        type=int,
        default=DEFAULT_SLEEP_MS,
        help=f"Pausa (ms) entre páginas OCR (por defecto {DEFAULT_SLEEP_MS})",
    )
    parser.add_argument(
        "--throttle-every",
        type=int,
        default=DEFAULT_THROTTLE_EVERY,
        help=f"Cada N PDFs aplicar pausa (por defecto {DEFAULT_THROTTLE_EVERY})",
    )
    parser.add_argument(
        "--throttle-ms",
        type=int,
        default=DEFAULT_THROTTLE_MS,
        help=f"Pausa (ms) cuando se cumple throttle-every (por defecto {DEFAULT_THROTTLE_MS})",
    )
    parser.add_argument(
        "--child-timeout-s",
        type=int,
        default=DEFAULT_CHILD_TIMEOUT_S,
        help=f"Timeout por PDF (por defecto {DEFAULT_CHILD_TIMEOUT_S})",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log), format="%(levelname)s: %(message)s"
    )

    input_dir = Path(args.input).expanduser().resolve()
    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(
            f"La carpeta de entrada no existe o no es una carpeta: {input_dir}"
        )
    output_xlsx = (
        Path(args.output).expanduser().resolve()
        if args.output
        else (input_dir / "facturas_datos_extraidos.xlsx")
    )

    pdfs = sorted([p for p in input_dir.glob("**/*.pdf")])
    if not pdfs:
        raise SystemExit(f"No se han encontrado PDFs en: {input_dir}")
    logging.info(f"Procesando {len(pdfs)} PDFs desde {input_dir}")

    rows: List[Row] = []
    count = 0
    for p in pdfs:
        count += 1
        try:
            recs = run_child_extract(
                p,
                dpi=max(72, int(args.dpi)),
                sleep_ms=max(0, int(args.sleep_ms)),
                poppler_path=args.poppler,
                tesseract_exe=args.tesseract,
                timeout_s=max(30, int(args.child_timeout_s)),
            )
            rows.extend(recs)
        except Exception as e:
            logging.error(f"Error en {p.name}: {e}")
        if (
            args.throttle_every
            and args.throttle_ms
            and (count % args.throttle_every == 0)
        ):
            import time

            time.sleep(args.throttle_ms / 1000.0)
        gc.collect()

    df = pd.DataFrame(
        rows,
        columns=[
            "fecha_factura",
            "numero_factura",
            "empresa",
            "CIF",
            "importe_base",
            "%IVA",
            "IVA",
            "importe_total",
            "Notas",
        ],
    )
    try:
        df["fecha_sort"] = pd.to_datetime(df["fecha_factura"], errors="coerce")
        df = df.sort_values(["fecha_sort", "numero_factura"]).drop(
            columns=["fecha_sort"]
        )
        df["fecha_factura"] = pd.to_datetime(df["fecha_factura"], errors="coerce")
    except Exception:
        pass

    df.to_excel(output_xlsx, index=False, engine="openpyxl")
    format_excel(output_xlsx)
    logging.info(f"Excel generado: {output_xlsx}")


if __name__ == "__main__":
    try:
        import multiprocessing as mp

        mp.freeze_support()
    except Exception:
        pass
    main()
